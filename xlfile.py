import csv
import math
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter  # Fixed: Added missing import


def load_heart_rate_data(file_path):
    """Reads a CSV file and returns a dictionary of {log_time: heart_rate_value}."""
    hr_dict = {}
    if not os.path.exists(file_path):
        return None

    with open(file_path, mode="r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)

        if not header:
            return hr_dict

        try:
            time_idx = header.index("log_time")
            hr_idx = -1
            for i, h in enumerate(header):
                cleaned = h.lower().replace("_", " ")
                if "heart" in cleaned or "hr" in cleaned or "rate" in cleaned:
                    hr_idx = i
                    break
            if hr_idx == -1:
                hr_idx = 1
        except ValueError:
            time_idx, hr_idx = 0, 1

        for row in reader:
            if not row or len(row) <= max(time_idx, hr_idx):
                continue
            time_str = row[time_idx].strip()
            hr_str = row[hr_idx].strip()

            try:
                hr_dict[time_str] = float(hr_str)
            except ValueError:
                continue

    return hr_dict


def calculate_metrics(ground_truth, predictions):
    """Calculates MAE, RMSE, and Percentage Accuracy based on MAPE."""
    absolute_errors = []
    squared_errors = []
    percentage_errors = []

    for time_key, gt_val in ground_truth.items():
        if time_key in predictions:
            pred_val = predictions[time_key]

            if gt_val == 0:
                continue

            error = pred_val - gt_val
            absolute_errors.append(abs(error))
            squared_errors.append(error**2)
            percentage_errors.append(abs(error) / gt_val)

    n_samples = len(absolute_errors)
    if n_samples == 0:
        return None, None, None, 0

    mae = sum(absolute_errors) / n_samples
    rmse = math.sqrt(sum(squared_errors) / n_samples)
    mape = sum(percentage_errors) / n_samples
    accuracy_pct = max(0.0, (1.0 - mape) * 100.0)

    return mae, rmse, accuracy_pct, n_samples


def build_excel_report(output_file, participant_results, global_stats):
    wb = openpyxl.Workbook()
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="595959")
    section_font = Font(name=font_family, size=12, bold=True, color="1F497D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, bold=False, color="000000")
    total_font = Font(name=font_family, size=11, bold=True, color="000000")
    kpi_num_font = Font(name=font_family, size=20, bold=True, color="1F497D")
    kpi_lbl_font = Font(name=font_family, size=9, bold=True, color="595959")

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    kpi_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    total_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom = Border(top=Side(border_style="thin", color="D9D9D9"), bottom=Side(border_style="double", color="1F497D"))

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # --- TAB 1: SUMMARY DASHBOARD ---
    ws1 = wb.active
    ws1.title = "Summary Dashboard"
    ws1.views.sheetView[0].showGridLines = True

    ws1["B2"] = "mmWave Radar Accuracy Benchmark Report"
    ws1["B2"].font = title_font
    ws1["B3"] = "Comparative Analysis against mam_sense ground truth references"
    ws1["B3"].font = subtitle_font

    ws1.merge_cells("B5:D5")
    ws1["B5"] = "OVERALL mmWAVE TI ACCURACY"
    ws1["B5"].font = kpi_lbl_font
    ws1["B5"].fill = kpi_fill
    ws1["B5"].alignment = align_center
    
    ws1.merge_cells("B6:D7")
    ws1["B6"] = global_stats["ti_acc"] / 100.0 if global_stats["ti_acc"] else "N/A"
    ws1["B6"].font = kpi_num_font
    ws1["B6"].fill = kpi_fill
    ws1["B6"].alignment = align_center
    ws1["B6"].number_format = "0.00%"

    ws1.merge_cells("F5:H5")
    ws1["F5"] = "OVERALL mmWAVE SS ACCURACY"
    ws1["F5"].font = kpi_lbl_font
    ws1["F5"].fill = kpi_fill
    ws1["F5"].alignment = align_center
    
    ws1.merge_cells("F6:H7")
    ws1["F6"] = global_stats["ss_acc"] / 100.0 if global_stats["ss_acc"] else "N/A"
    ws1["F6"].font = kpi_num_font
    ws1["F6"].fill = kpi_fill
    ws1["F6"].alignment = align_center
    ws1["F6"].number_format = "0.00%"

    for r in range(5, 8):
        for c in range(2, 5):
            ws1.cell(row=r, column=c).border = thin_border
        for c in range(6, 9):
            ws1.cell(row=r, column=c).border = thin_border

    ws1["B9"] = "Global Performance Metrics Summary"
    ws1["B9"].font = section_font
    
    headers_summary = ["Sensor Module", "Paired Samples", "Mean Absolute Error (MAE)", "Root Mean Squared Error (RMSE)"]
    for col_idx, text in enumerate(headers_summary, start=2):
        cell = ws1.cell(row=10, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    summary_data = [
        ["mmWave TI Module", global_stats["ti_samples"], global_stats["ti_mae"], global_stats["ti_rmse"]],
        ["mmWave SS Module", global_stats["ss_samples"], global_stats["ss_mae"], global_stats["ss_rmse"]]
    ]

    for row_idx, row_data in enumerate(summary_data, start=11):
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 2:
                cell.alignment = align_left
            elif col_idx == 3:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            else:
                cell.alignment = align_right
                cell.number_format = "0.000"

    # --- TAB 2: PARTICIPANT BREAKDOWN ---
    ws2 = wb.create_sheet(title="Participant Breakdown")
    ws2.views.sheetView[0].showGridLines = True

    ws2["B2"] = "Granular Evaluation Matrix per Participant Folder"
    ws2["B2"].font = title_font

    ws2.merge_cells("B4:B5")
    ws2["B4"] = "Participant ID"
    ws2.merge_cells("C4:F4")
    ws2["C4"] = "Texas Instruments Module (mmWave TI)"
    ws2.merge_cells("G4:J4")
    ws2["G4"] = "Silicon Labs Module (mmWave SS)"

    sub_headers = [
        "Accuracy %", "MAE (BPM)", "RMSE (BPM)", "Samples",
        "Accuracy %", "MAE (BPM)", "RMSE (BPM)", "Samples"
    ]
    for idx, sh in enumerate(sub_headers, start=3):
        ws2.cell(row=5, column=idx, value=sh)

    for c in range(2, 11):
        top_cell = ws2.cell(row=4, column=c)
        top_cell.font = header_font
        top_cell.fill = header_fill
        top_cell.alignment = align_center
        top_cell.border = thin_border

        sub_cell = ws2.cell(row=5, column=c)
        sub_cell.font = Font(name=font_family, size=10, bold=True, color="1F497D")
        sub_cell.fill = sub_header_fill
        sub_cell.alignment = align_center
        sub_cell.border = thin_border

    current_row = 6
    for pt_id, metrics in sorted(participant_results.items()):
        row_values = [
            pt_id,
            metrics["ti_acc"] / 100.0 if metrics["ti_acc"] else "",
            metrics["ti_mae"], metrics["ti_rmse"], metrics["ti_samples"] if metrics["ti_samples"] else "",
            metrics["ss_acc"] / 100.0 if metrics["ss_acc"] else "",
            metrics["ss_mae"], metrics["ss_rmse"], metrics["ss_samples"] if metrics["ss_samples"] else ""
        ]
        
        is_zebra = (current_row % 2 == 0)
        for col_idx, val in enumerate(row_values, start=2):
            cell = ws2.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if is_zebra:
                cell.fill = zebra_fill
            
            if col_idx == 2:
                cell.alignment = align_left
            elif col_idx in [3, 7]:
                cell.alignment = align_right
                cell.number_format = "0.00%"
            elif col_idx in [4, 5, 8, 9]:
                cell.alignment = align_right
                cell.number_format = "0.00"
            elif col_idx in [6, 10]:
                cell.alignment = align_right
                cell.number_format = "#,##0"
        current_row += 1

    avg_row = current_row
    ws2.cell(row=avg_row, column=2, value="Averages / Totals").font = total_font
    ws2.cell(row=avg_row, column=2).fill = total_fill
    ws2.cell(row=avg_row, column=2).border = double_bottom

    formula_mappings = {
        3: f"=AVERAGE(C6:C{avg_row-1})", 4: f"=AVERAGE(D6:D{avg_row-1})",
        5: f"=AVERAGE(E6:E{avg_row-1})", 6: f"=SUM(F6:F{avg_row-1})",
        7: f"=AVERAGE(G6:G{avg_row-1})", 8: f"=AVERAGE(H6:H{avg_row-1})",
        9: f"=AVERAGE(I6:I{avg_row-1})", 10: f"=SUM(J6:J{avg_row-1})"
    }

    for c_idx, formula in formula_mappings.items():
        cell = ws2.cell(row=avg_row, column=c_idx, value=formula)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = double_bottom
        if c_idx in [3, 7]:
            cell.number_format = "0.00%"
            cell.alignment = align_right
        elif c_idx in [4, 5, 8, 9]:
            cell.number_format = "0.00"
            cell.alignment = align_right
        else:
            cell.number_format = "#,##0"
            cell.alignment = align_right

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(name=font_family, size=11, color="9C0006")
    alert_rule = CellIsRule(operator="lessThan", formula=["0.85"], stopIfTrue=True, fill=red_fill, font=red_font)
    ws2.conditional_formatting.add(f"C6:C{avg_row-1}", alert_rule)
    ws2.conditional_formatting.add(f"G6:G{avg_row-1}", alert_rule)

    for sheet in [ws1, ws2]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    if not val_str.startswith("="):
                        max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_file)


def execute_evaluation_to_spreadsheet():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_base_dir = os.path.join(base_dir, "output")
    excel_output_path = os.path.join(base_dir, "Heart_Rate_Evaluation_Report.xlsx")

    if not os.path.exists(output_base_dir):
        print(f"Error: Output folder not found at {output_base_dir}")
        return

    global_ti_errors, global_ti_squared, global_ti_pct_errs = [], [], []
    global_ss_errors, global_ss_squared, global_ss_pct_errs = [], [], []
    participant_results = {}

    # --- NEW: PRINT INDIVIDUAL PARTICIPANT DATA IN TERMINAL ---
    print("=" * 80)
    print(" INDIVIDUAL PARTICIPANT ANALYTICAL DATA BREAKDOWN")
    print("=" * 80)

    for folder_name in sorted(os.listdir(output_base_dir)):
        folder_path = os.path.join(output_base_dir, folder_name)
        if not os.path.isdir(folder_path):
            continue

        mam_file = os.path.join(folder_path, "mam_sense.csv")
        ti_file = os.path.join(folder_path, "mmwave_ti_heart_rate.csv")
        ss_file = os.path.join(folder_path, "mmwave_ss_heart.csv")

        mam_data = load_heart_rate_data(mam_file)
        if not mam_data:
            continue

        # Process mmWave TI Module
        ti_data = load_heart_rate_data(ti_file)
        mae_ti, rmse_ti, acc_ti, samples_ti = None, None, None, 0
        if ti_data:
            mae_ti, rmse_ti, acc_ti, samples_ti = calculate_metrics(mam_data, ti_data)
            for tk, gt in mam_data.items():
                if tk in ti_data and gt != 0:
                    err = ti_data[tk] - gt
                    global_ti_errors.append(abs(err))
                    global_ti_squared.append(err**2)
                    global_ti_pct_errs.append(abs(err) / gt)

        # Process mmWave SS Module
        ss_data = load_heart_rate_data(ss_file)
        mae_ss, rmse_ss, acc_ss, samples_ss = None, None, None, 0
        if ss_data:
            mae_ss, rmse_ss, acc_ss, samples_ss = calculate_metrics(mam_data, ss_data)
            for tk, gt in mam_data.items():
                if tk in ss_data and gt != 0:
                    err = ss_data[tk] - gt
                    global_ss_errors.append(abs(err))
                    global_ss_squared.append(err**2)
                    global_ss_pct_errs.append(abs(err) / gt)

        participant_results[folder_name] = {
            "ti_acc": acc_ti, "ti_mae": mae_ti, "ti_rmse": rmse_ti, "ti_samples": samples_ti,
            "ss_acc": acc_ss, "ss_mae": mae_ss, "ss_rmse": rmse_ss, "ss_samples": samples_ss
        }

        # Terminal Print Display
        print(f"\n[ Participant: {folder_name} ]")
        if samples_ti > 0:
            print(f"  -> mmWave TI: Accuracy = {acc_ti:6.2f}% | MAE = {mae_ti:5.2f} BPM | RMSE = {rmse_ti:5.2f} BPM ({samples_ti} samples)")
        else:
            print("  -> mmWave TI: No valid paired data or file missing.")

        if samples_ss > 0:
            print(f"  -> mmWave SS: Accuracy = {acc_ss:6.2f}% | MAE = {mae_ss:5.2f} BPM | RMSE = {rmse_ss:5.2f} BPM ({samples_ss} samples)")
        else:
            print("  -> mmWave SS: No valid paired data or file missing.")

    # Calculate global aggregates
    global_stats = {
        "ti_samples": len(global_ti_errors),
        "ti_mae": sum(global_ti_errors) / len(global_ti_errors) if global_ti_errors else None,
        "ti_rmse": math.sqrt(sum(global_ti_squared) / len(global_ti_squared)) if global_ti_squared else None,
        "ti_acc": max(0.0, (1.0 - (sum(global_ti_pct_errs) / len(global_ti_pct_errs))) * 100.0) if global_ti_pct_errs else None,
        
        "ss_samples": len(global_ss_errors),
        "ss_mae": sum(global_ss_errors) / len(global_ss_errors) if global_ss_errors else None,
        "ss_rmse": math.sqrt(sum(global_ss_squared) / len(global_ss_squared)) if global_ss_squared else None,
        "ss_acc": max(0.0, (1.0 - (sum(global_ss_pct_errs) / len(global_ss_pct_errs))) * 100.0) if global_ss_pct_errs else None,
    }

    # Print Global Analytics Summary to Terminal
    print("\n" + "=" * 80)
    print(" OVERALL SYSTEM SUMMARY METRICS")
    print("=" * 80)
    if global_stats["ti_acc"]:
        print(f"Overall mmWave TI Module Performance:")
        print(f"  Total Data Points: {global_stats['ti_samples']}")
        print(f"  Accuracy (MAPE):   {global_stats['ti_acc']:.2f}%")
        print(f"  MAE:               {global_stats['ti_mae']:.3f} BPM")
        print(f"  RMSE:              {global_stats['ti_rmse']:.3f} BPM")
    print("-" * 80)
    if global_stats["ss_acc"]:
        print(f"Overall mmWave SS Module Performance:")
        print(f"  Total Data Points: {global_stats['ss_samples']}")
        print(f"  Accuracy (MAPE):   {global_stats['ss_acc']:.2f}%")
        print(f"  MAE:               {global_stats['ss_mae']:.3f} BPM")
        print(f"  RMSE:              {global_stats['ss_rmse']:.3f} BPM")
    print("=" * 80 + "\n")

    # Build spreadsheet
    build_excel_report(excel_output_path, participant_results, global_stats)
    print(f"Spreadsheet evaluation successfully compiled! -> {excel_output_path}")


if __name__ == "__main__":
    execute_evaluation_to_spreadsheet()